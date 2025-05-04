# logger.py

import openpyxl

def write_headers(sheet, headers, start_row=1, start_col=1):
    """
    כותב רשימת כותרות בגליון Excel, החל מתא (start_row, start_col).
    """
    for idx, header in enumerate(headers):
        sheet.cell(row=start_row, column=start_col + idx, value=header)

def write_data(sheet, data, start_row=2, start_col=1):
    """
    כותב רשימת שורות (data) בגליון Excel, החל מ-(start_row, start_col).
    כל שורה ב‑data היא iterable של ערכים לכל תא.
    """
    for r_idx, row in enumerate(data, start=start_row):
        for c_idx, val in enumerate(row, start=start_col):
            sheet.cell(row=r_idx, column=c_idx, value=val)

def find_next_available_row(sheet, start_col=1):
    """
    מחזיר את מספר השורה הבא שבו התא ב‑start_col ריק.
    """
    row = 1
    while sheet.cell(row=row, column=start_col).value is not None:
        row += 1
    return row
