# main.py

import openpyxl
from datetime import date
from assign import process_row
from calculate import perform_calculation
from logger import write_headers, write_data
import refdata

def main():
    # 0. טען מפות ההנחות והתמותה דרך refdata
    discount_map  = refdata.load_assumptions("data2.xlsx")
    mortality_map = refdata.load_mortality("תמותה.xlsx")

    # 1. טען את קובץ הנתונים (data2.xlsx)
    wb = openpyxl.load_workbook("data2.xlsx", data_only=True)
    data_sheet = wb["data"]

    # 2. קבע תאריך חיתוך לחישוב
    calculation_date = date(2024, 12, 31)

    # 3. הכנה של גליון התוצאות
    result_wb    = openpyxl.Workbook()
    result_sheet = result_wb.active
    result_sheet.title = "התחייבות"
    headers = ["emp_id", "שם מלא", "תחייבות"]
    write_headers(result_sheet, headers, start_row=1, start_col=1)

    # 4. עיבוד כל שורה בגליון data
    results = []
    for row_idx in range(3, data_sheet.max_row + 1):
        row_data = process_row(data_sheet, row_idx)
        if not row_data:
            continue

        (
            emp_id, birthdate, person_age, final_salary, full_name,
            start_date, leave_date, article14_date,
            article14_rate_cell, asset_value_cell, gender_cell
        ) = row_data

        # 5. עובדים שעזבו ב-2024 => התחייבות 0
        if leave_date and leave_date.year == 2024:
            liability = 0.0
        else:
            salary_growth = 0.04
            liability = perform_calculation(
                start_date=start_date,
                leave_date=leave_date,
                article14_date=article14_date,
                salary_growth=salary_growth,
                final_salary=final_salary,
                person_age=person_age,
                person_gender=gender_cell.value,
                calculation_date=calculation_date,
                discount_map=discount_map,
                mortality_map=mortality_map
            )

        results.append([emp_id, full_name, liability])

    # 6. כתיבת כל התוצאות בגליון
    write_data(result_sheet, results, start_row=2, start_col=1)

    # 7. שמירת קובץ ה־Excel
    result_wb.save("results.xlsx")
    print("Saved results.xlsx")

if __name__ == "__main__":
    main()
