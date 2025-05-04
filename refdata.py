# refdata.py

import openpyxl
from datetime import date

def load_employee_data(path="data2.xlsx"):
    """
    טוען את גליון הנתונים 'data' ממסלול הנתונים ומחזיר רשימת dict לכל עובד עם השדות הדרושים:
    emp_id (מספר שורה-1), full_name, gender, birthdate, start_date, salary, article14_date, article14_rate, leave_date
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = wb["data"]
    employees = []

    for row_idx, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
        # מיפוי עמודות לפי data2.xlsx:
        first_name     = row[1]
        last_name      = row[2]
        gender         = row[3]
        birthdate      = row[4]
        start_date     = row[5]
        salary         = row[6]
        article14_date = row[7]
        article14_rate = row[8]
        leave_date     = row[11]

        emp = {
            "emp_id":         row_idx - 1,             # מזהה פנימי
            "full_name":      f"{first_name} {last_name}",
            "gender":         gender,
            "birthdate":      birthdate,
            "start_date":     start_date,
            "salary":         salary,
            "article14_date": article14_date,
            "article14_rate": article14_rate,
            "leave_date":     leave_date,
        }
        employees.append(emp)
    return employees

def load_assumptions(path="data2.xlsx"):
    """
    טוען את גליון 'הנחות' ומחזיר dict של שנים→שיעור היוון.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = wb["הנחות"]
    discount_map = {}
    for row in sheet.iter_rows(min_row=5, min_col=1, max_col=2, values_only=True):
        work_years, rate = row
        if work_years is None:
            continue
        discount_map[int(work_years)] = rate
    return discount_map

def load_mortality(path="תמותה.xlsx"):
    """
    טוען את שני הגליונות 'גברים' ו'נשים' ומחזיר dict:
    { gender_key: { age: q_x, ... }, ... }
    כאשר gender_key הוא 0 עבור גברים, 1 עבור נשים.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    mortality_map = {}
    for gender_key, sheet_name in [(0, "גברים"), (1, "נשים")]:
        sheet = wb[sheet_name]
        age_map = {}
        for row in sheet.iter_rows(min_row=3, values_only=True):
            age = row[1]
            q_x  = row[5]
            if age is None:
                continue
            age_map[int(age)] = q_x
        mortality_map[gender_key] = age_map
    return mortality_map

def get_calculation_date():
    """מחזיר את תאריך החיתוך לחישוב על פי ההנחיות."""
    return date(2024, 12, 31)
